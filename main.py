import re
import os
import pandas as pd
import os
from decouple import config

from custom_chains import vision_model, product_info_parser
from utils import download_pdf_and_convert_to_images, get_column_value,get_column_value_from_ws
from models import product_output_structure
from concurrent.futures import ThreadPoolExecutor, as_completed
import openpyxl

# Set verbose
# globals.set_debug(True)

os.environ["OPENAI_API_KEY"] = config("OPENAI_KEY")


def prepare_prompt(product_name, row_data, pdf_text, pdf_tables_data):
    prompt = f"""
    You are intelligent agent who is expert in extracting structured data from raw text.
    You specilize in extracting title/product name, product specifications and product features list from a text data of spec-sheet pdf and images of pdf pages.

    Here is data of product extracted from excel sheet = {row_data}
    Here is text extracted from spec-sheet of product called {product_name} = {pdf_text}
    Also here is the tablur data extracted from same pdf = {pdf_tables_data}
    

    You need to analyze it hard and extract title/product name, specifications list and  list of features into json of specified format.

    Remeber that the data is very unstructured, text is very ambigious, so you need to analyze hard 
    and extract title, specifications & features. If tabular data is extracted from pdf, then it can be used for specifications.
    This data will be presented on E-Commerce website.
    Images of pages of pdf file are also provided, analyze those images and use them as well.
    The features should be an html bullet list. and you must not miss any feature but also do not put anything that is not feature.
    Also do not mix specifications and features.
    The specifications should be a 2 column html table of specifications without styles or headings but there must be a tbody tag in table.
    Also UPC code can be ignored for the specifications tables. Also make sure that there is no duplicate information in specifications table.
    Description of product should be generated in p tag using product data especiall Descriptio Column, features and specs... Generate description that will be perfect for showcasing on Ecommerce sites.
    Description must be other worldly.... like user must buy the item after reading description...
    The warranty should be in the following format: 1:Labor, 2:Parts if it says 1 year labor and 2 year parts else if it says 1 year labor and parts then you should put 1:Labor, 1:Parts. I hope you understand the format of warranty.
    Also make sure that warranty is in years , warranty in months should be divided by 12 to get warranty in years 40:Labor, 80:Parts -> 3.3:Labor, 6.6:Parts. if value for warranty is big which can mean that it is month, so convert it into years.
    Specifications extracted, must really be specifications, and features must really be features of product.
    Do not put random text into features list.
    The comma seperated certifications should be gramatically correct, e.g capitalization. You can also try to find certifications from Column "DESCRIPTION" present in data of product extracted from excel sheet
    We can ignore FOR SALE IN NORTH AMERICA whenever it says FOR SALE IN NORTH AMERICA in column "DESCRIPTION" in data of product extracted from excel sheet.
    And "ETL-Sanitation" should be "ETLSAN" and "cETLus" should be "CELTUS" if present for certifications.
    Also I must mention that although the povided text from pdf can be different 
    but I have noticed that points starting with • or * is a feature... • or * is present before feature but sometimes * is also there for spec so think carefully.
    Also remeber to not use characters that are not human readable e.g other than English.
    Especially â€/â€ means inches, Â° means degree of temprature... and  so be careful of symbols that are not human readable.
    Also make sure that you only extract data from pdf for the same product whose data of product is given from excel sheet.
    Everything should be cleaned up and human readable.
    """
    return prompt


def get_product_informations(
    product_name, row_data, pdf_text, images_path: str, pdf_tables_data
) -> dict:
    prompt = prepare_prompt(product_name, row_data, pdf_text, pdf_tables_data)
    # vision_chain = load_image_chain | vision_model | product_info_parser
    vision_chain = vision_model | product_info_parser
    return vision_chain.invoke({"images_path": f"{images_path}", "prompt": prompt})


def sanitize_data(data):
    # Remove characters that are illegal in Excel
    return re.sub(r"[\x00-\x1F\x7F-\x9F]", "", data)


def process_row(index, row):
    try:
        print("\nGoing to process row at index ", index)
        pdf_file_url = row["Spec sheet"]
        if not pd.notna(pdf_file_url):
            return index, None

        pdf_info = download_pdf_and_convert_to_images(pdf_file_url)

        product_info = get_product_informations(
            get_column_value(row, "PRODUCT"),
            row.to_json(),
            pdf_info["pdf_text"],
            pdf_info["images_path"],
            pdf_info["tables_data"],
        )
        print("\nGot Product Info for row at index ", index)
        print(product_info)
        sku = get_column_value(row, "SKU")
        sku = "" if sku is None else sku
        product_name = (
            f"{get_column_value(row, 'Brand')} {sku} {get_column_value(row, 'PRODUCT')}"
        )
        return index, [
            product_name,
            sanitize_data(product_info["product_description"]),
            sanitize_data(pdf_info["pdf_text"]),
            sanitize_data(product_info["product_features"]),
            sanitize_data(product_info["product_specifications"]),
            sanitize_data(product_info["product_certifications"]),
            sanitize_data(product_info["product_warranty"]),
            "",
        ]
    except Exception as ex:
        print(ex)
        return index, ["", "", "", "", "", "", "", str(ex)]


def process_excel_sheet(excel_path, row_index=None):
    df = pd.read_excel(excel_path)
    df.columns = df.columns.str.strip()
    print(df.columns)
    print("==============")
    print(df.head())
    results_df = product_output_structure

    with ThreadPoolExecutor(max_workers=3) as executor:
        futures = [
            executor.submit(process_row, index, row)
            for index, row in df.head(3).iterrows()
            if row_index is None or index == row_index
        ]

        for future in as_completed(futures):
            index, result = future.result()
            if result:
                results_df.loc[index] = result

    # Save to a new Excel file
    output_path = "Resources/output.xlsx"
    try:
        results_df.to_excel(output_path, index=False)
        print(f"Data successfully saved to {output_path}")
    except Exception as e:
        print(f"Error saving data to Excel: {e}")
    results_df.to_excel("Resources/output.xlsx", index=False)

    # Load the original workbook and update it
    wb = openpyxl.load_workbook(excel_path)
    ws = wb.active

    # Strip column names in both sheets
    ws1_columns = [
        cell.value.strip() if cell.value is not None else "" for cell in ws[1]
    ]

    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        sku = get_column_value_from_ws(row, 'SKU', ws)
        sku = "" if sku is None else sku
        product_name = f"{get_column_value_from_ws(row, 'Brand', ws)} {sku} {get_column_value_from_ws(row, 'PRODUCT', ws)}"

        # Find the matching row in results_df
        for result_idx, result_row in results_df.iterrows():
            if result_row['S2.Product Name'] == product_name:
                ws.cell(row=row[0].row, column=ws1_columns.index('S2.Product Name') + 1, value=result_row['S2.Product Name'])
                ws.cell(row=row[0].row, column=ws1_columns.index('S2.Description') + 1, value=result_row['S2.Description'])
                ws.cell(row=row[0].row, column=ws1_columns.index('S2.Pdf_Text') + 1, value=result_row['S2.PDF Text'])
                ws.cell(row=row[0].row, column=ws1_columns.index('S2.Features') + 1, value=result_row['S2.Features'])
                ws.cell(row=row[0].row, column=ws1_columns.index('S2.Specifications') + 1, value=result_row['S2.Specifications'])
                ws.cell(row=row[0].row, column=ws1_columns.index('S2.Certifications') + 1, value=result_row['S2.Certifications'])
                ws.cell(row=row[0].row, column=ws1_columns.index('S2.Warranty') + 1, value=result_row['S2.Warranty'])
                break

    # Save the updated original workbook
    updated_path = "Resources/updated_sheet1.xlsx"
    wb.save(updated_path)
    print(f"Original data successfully updated and saved to {updated_path}")


if __name__ == "__main__":
    excel_path = "Resources/MVP Group MAP-MCOP  MAY 1, 2024-Revised.xlsx"
    process_excel_sheet(excel_path)
